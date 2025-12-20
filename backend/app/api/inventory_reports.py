"""
API para geração de relatórios de inventário
PDF e Excel com resumo e detalhamento
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import Optional
from datetime import datetime
import io
import logging

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import cm, mm

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from app.core.database import get_db
from app.core.auth import get_current_user
from app.models import (
    User,
    InventorySession,
    InventoryExpectedAsset,
    InventoryReadAsset,
    AssetCategory,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/inventory/reports", tags=["inventory-reports"])


def calculate_session_statistics(db: Session, session_id: int) -> dict:
    """Calcula estatísticas da sessão"""
    total_expected = db.query(func.count(InventoryExpectedAsset.id)).filter(
        InventoryExpectedAsset.session_id == session_id
    ).scalar() or 0

    category_counts = db.query(
        InventoryReadAsset.category,
        func.count(InventoryReadAsset.id)
    ).filter(
        InventoryReadAsset.session_id == session_id
    ).group_by(InventoryReadAsset.category).all()

    stats = {
        "total_expected": total_expected,
        "total_found": 0,
        "total_not_found": 0,
        "total_unregistered": 0,
        "total_written_off": 0,
        "completion_percentage": 0.0
    }

    for category, count in category_counts:
        if category == AssetCategory.FOUND.value:
            stats["total_found"] = count
        elif category == AssetCategory.NOT_FOUND.value:
            stats["total_not_found"] = count
        elif category == AssetCategory.UNREGISTERED.value:
            stats["total_unregistered"] = count
        elif category == AssetCategory.WRITTEN_OFF.value:
            stats["total_written_off"] = count

    if total_expected > 0:
        verified = stats["total_found"] + stats["total_not_found"] + stats["total_written_off"]
        stats["completion_percentage"] = round((verified / total_expected) * 100, 2)

    return stats


@router.get("/{session_id}/pdf")
async def generate_pdf_report(
    session_id: int,
    include_details: bool = Query(True, description="Incluir lista detalhada de bens"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Gera relatório PDF da sessão de inventário.
    Inclui resumo estatístico e opcionalmente lista detalhada de bens.
    """
    session = db.query(InventorySession).options(
        joinedload(InventorySession.project),
        joinedload(InventorySession.ug),
        joinedload(InventorySession.ul),
        joinedload(InventorySession.ua)
    ).filter(InventorySession.id == session_id).first()

    if not session:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")

    stats = calculate_session_statistics(db, session_id)

    # Criar PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12
    )
    normal_style = styles['Normal']

    elements = []

    # Título
    elements.append(Paragraph("Relatório de Inventário Patrimonial", title_style))
    elements.append(Spacer(1, 20))

    # Informações da Sessão
    elements.append(Paragraph("Informações da Sessão", subtitle_style))

    session_info = [
        ["Código", session.code],
        ["Nome", session.name or "-"],
        ["Status", {
            'draft': 'Rascunho',
            'in_progress': 'Em Andamento',
            'paused': 'Pausada',
            'completed': 'Concluída',
            'cancelled': 'Cancelada'
        }.get(session.status, session.status)],
        ["Projeto", session.project.nome if session.project else "-"],
        ["UG", f"{session.ug.code} - {session.ug.name}" if session.ug else "-"],
        ["UL", f"{session.ul.code} - {session.ul.name}" if session.ul else "-"],
        ["Data Criação", session.created_at.strftime("%d/%m/%Y %H:%M") if session.created_at else "-"],
        ["Data Início", session.started_at.strftime("%d/%m/%Y %H:%M") if session.started_at else "-"],
        ["Data Conclusão", session.completed_at.strftime("%d/%m/%Y %H:%M") if session.completed_at else "-"],
    ]

    t = Table(session_info, colWidths=[5*cm, 12*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # Estatísticas
    elements.append(Paragraph("Resumo Estatístico", subtitle_style))

    stats_data = [
        ["Indicador", "Quantidade", "Percentual"],
        ["Total Esperado", str(stats["total_expected"]), "100%"],
        ["Encontrados", str(stats["total_found"]), f"{(stats['total_found']/stats['total_expected']*100 if stats['total_expected'] > 0 else 0):.1f}%"],
        ["Não Encontrados", str(stats["total_not_found"]), f"{(stats['total_not_found']/stats['total_expected']*100 if stats['total_expected'] > 0 else 0):.1f}%"],
        ["Não Cadastrados", str(stats["total_unregistered"]), "-"],
        ["Baixados", str(stats["total_written_off"]), f"{(stats['total_written_off']/stats['total_expected']*100 if stats['total_expected'] > 0 else 0):.1f}%"],
        ["Conclusão", f"{stats['completion_percentage']:.1f}%", ""],
    ]

    t = Table(stats_data, colWidths=[7*cm, 5*cm, 5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#dcfce7')),  # Encontrados - verde claro
        ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fee2e2')),  # Não encontrados - vermelho claro
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    # Lista detalhada de bens
    if include_details:
        elements.append(PageBreak())
        elements.append(Paragraph("Lista Detalhada de Bens", subtitle_style))

        # Buscar bens esperados com leituras
        expected_assets = db.query(InventoryExpectedAsset).filter(
            InventoryExpectedAsset.session_id == session_id
        ).order_by(InventoryExpectedAsset.asset_code).all()

        asset_ids = [a.id for a in expected_assets]
        readings = db.query(InventoryReadAsset).filter(
            InventoryReadAsset.expected_asset_id.in_(asset_ids)
        ).all()
        readings_map = {r.expected_asset_id: r for r in readings}

        # Tabela de bens
        assets_data = [["Código", "Descrição", "RFID", "Status", "Data Leitura"]]

        for asset in expected_assets:
            reading = readings_map.get(asset.id)
            status = "Pendente"
            read_at = "-"

            if reading:
                status_map = {
                    'found': 'Encontrado',
                    'not_found': 'Não Encontrado',
                    'unregistered': 'Não Cadastrado',
                    'written_off': 'Baixado'
                }
                status = status_map.get(reading.category, reading.category)
                read_at = reading.read_at.strftime("%d/%m/%Y %H:%M") if reading.read_at else "-"
            elif asset.is_written_off:
                status = "Baixado"

            # Truncar descrição se muito longa
            desc = asset.description or "-"
            if len(desc) > 40:
                desc = desc[:37] + "..."

            assets_data.append([
                asset.asset_code,
                desc,
                asset.rfid_code or "-",
                status,
                read_at
            ])

        # Dividir em múltiplas tabelas se necessário (máx 40 linhas por página)
        chunk_size = 40
        for i in range(0, len(assets_data), chunk_size):
            chunk = assets_data[i:i+chunk_size]
            if i > 0:
                chunk.insert(0, assets_data[0])  # Adicionar cabeçalho

            t = Table(chunk, colWidths=[3*cm, 6*cm, 3*cm, 2.5*cm, 3*cm])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('PADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)

            if i + chunk_size < len(assets_data):
                elements.append(PageBreak())

    # Rodapé
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} por {current_user.nome}",
        ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.grey)
    ))

    doc.build(elements)
    buffer.seek(0)

    filename = f"inventario_{session.code}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{session_id}/excel")
async def generate_excel_report(
    session_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Gera relatório Excel da sessão de inventário.
    Inclui abas de resumo e lista detalhada de bens.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Biblioteca openpyxl não disponível")

    session = db.query(InventorySession).options(
        joinedload(InventorySession.project),
        joinedload(InventorySession.ug),
        joinedload(InventorySession.ul),
        joinedload(InventorySession.ua)
    ).filter(InventorySession.id == session_id).first()

    if not session:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")

    stats = calculate_session_statistics(db, session_id)

    # Criar workbook
    wb = openpyxl.Workbook()

    # Estilos
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=14, bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aba 1: Resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    # Título
    ws_resumo['A1'] = "Relatório de Inventário Patrimonial"
    ws_resumo['A1'].font = Font(size=16, bold=True)
    ws_resumo.merge_cells('A1:D1')

    # Informações da sessão
    ws_resumo['A3'] = "Informações da Sessão"
    ws_resumo['A3'].font = title_font

    info_data = [
        ("Código", session.code),
        ("Nome", session.name or "-"),
        ("Status", {
            'draft': 'Rascunho',
            'in_progress': 'Em Andamento',
            'paused': 'Pausada',
            'completed': 'Concluída',
            'cancelled': 'Cancelada'
        }.get(session.status, session.status)),
        ("Projeto", session.project.nome if session.project else "-"),
        ("UG", f"{session.ug.code} - {session.ug.name}" if session.ug else "-"),
        ("UL", f"{session.ul.code} - {session.ul.name}" if session.ul else "-"),
        ("Data Criação", session.created_at.strftime("%d/%m/%Y %H:%M") if session.created_at else "-"),
        ("Data Início", session.started_at.strftime("%d/%m/%Y %H:%M") if session.started_at else "-"),
        ("Data Conclusão", session.completed_at.strftime("%d/%m/%Y %H:%M") if session.completed_at else "-"),
    ]

    for i, (label, value) in enumerate(info_data, start=4):
        ws_resumo[f'A{i}'] = label
        ws_resumo[f'A{i}'].font = Font(bold=True)
        ws_resumo[f'B{i}'] = value

    # Estatísticas
    row = len(info_data) + 6
    ws_resumo[f'A{row}'] = "Resumo Estatístico"
    ws_resumo[f'A{row}'].font = title_font

    row += 2
    stats_headers = ["Indicador", "Quantidade", "Percentual"]
    for col, header in enumerate(stats_headers, start=1):
        cell = ws_resumo.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    stats_data = [
        ("Total Esperado", stats["total_expected"], "100%"),
        ("Encontrados", stats["total_found"], f"{(stats['total_found']/stats['total_expected']*100 if stats['total_expected'] > 0 else 0):.1f}%"),
        ("Não Encontrados", stats["total_not_found"], f"{(stats['total_not_found']/stats['total_expected']*100 if stats['total_expected'] > 0 else 0):.1f}%"),
        ("Não Cadastrados", stats["total_unregistered"], "-"),
        ("Baixados", stats["total_written_off"], f"{(stats['total_written_off']/stats['total_expected']*100 if stats['total_expected'] > 0 else 0):.1f}%"),
        ("Conclusão", f"{stats['completion_percentage']:.1f}%", ""),
    ]

    for i, (label, qty, pct) in enumerate(stats_data, start=1):
        for col, value in enumerate([label, qty, pct], start=1):
            cell = ws_resumo.cell(row=row+i, column=col, value=value)
            cell.border = thin_border
            if col > 1:
                cell.alignment = Alignment(horizontal='center')

    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 20
    ws_resumo.column_dimensions['B'].width = 40
    ws_resumo.column_dimensions['C'].width = 15

    # Aba 2: Bens Detalhados
    ws_bens = wb.create_sheet("Bens Detalhados")

    # Cabeçalho
    headers = ["Código do Bem", "Descrição", "Tag RFID", "Código Barras", "UL Esperada", "Status", "Método Leitura", "Data Leitura", "Estado Físico"]
    for col, header in enumerate(headers, start=1):
        cell = ws_bens.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Buscar bens
    expected_assets = db.query(InventoryExpectedAsset).filter(
        InventoryExpectedAsset.session_id == session_id
    ).order_by(InventoryExpectedAsset.asset_code).all()

    asset_ids = [a.id for a in expected_assets]
    readings = db.query(InventoryReadAsset).filter(
        InventoryReadAsset.expected_asset_id.in_(asset_ids)
    ).all()
    readings_map = {r.expected_asset_id: r for r in readings}

    # Dados dos bens
    for i, asset in enumerate(expected_assets, start=2):
        reading = readings_map.get(asset.id)

        status = "Pendente"
        read_method = "-"
        read_at = "-"
        physical_condition = "-"

        if reading:
            status_map = {
                'found': 'Encontrado',
                'not_found': 'Não Encontrado',
                'unregistered': 'Não Cadastrado',
                'written_off': 'Baixado'
            }
            status = status_map.get(reading.category, reading.category)
            read_method = reading.read_method or "-"
            read_at = reading.read_at.strftime("%d/%m/%Y %H:%M") if reading.read_at else "-"
            physical_condition = reading.physical_condition or "-"
        elif asset.is_written_off:
            status = "Baixado"

        row_data = [
            asset.asset_code,
            asset.description or "-",
            asset.rfid_code or "-",
            asset.barcode or "-",
            asset.expected_ul_code or "-",
            status,
            read_method,
            read_at,
            physical_condition
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws_bens.cell(row=i, column=col, value=value)
            cell.border = thin_border

        # Colorir status
        status_cell = ws_bens.cell(row=i, column=6)
        if status == "Encontrado":
            status_cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
        elif status == "Não Encontrado":
            status_cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
        elif status == "Não Cadastrado":
            status_cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")

    # Ajustar largura das colunas
    ws_bens.column_dimensions['A'].width = 15
    ws_bens.column_dimensions['B'].width = 40
    ws_bens.column_dimensions['C'].width = 20
    ws_bens.column_dimensions['D'].width = 15
    ws_bens.column_dimensions['E'].width = 15
    ws_bens.column_dimensions['F'].width = 15
    ws_bens.column_dimensions['G'].width = 15
    ws_bens.column_dimensions['H'].width = 18
    ws_bens.column_dimensions['I'].width = 15

    # Aba 3: Não Cadastrados (se houver)
    unregistered = db.query(InventoryReadAsset).filter(
        InventoryReadAsset.session_id == session_id,
        InventoryReadAsset.category == AssetCategory.UNREGISTERED.value
    ).all()

    if unregistered:
        ws_unreg = wb.create_sheet("Não Cadastrados")

        headers = ["Identificador", "Método Leitura", "Data Leitura"]
        for col, header in enumerate(headers, start=1):
            cell = ws_unreg.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border

        for i, reading in enumerate(unregistered, start=2):
            row_data = [
                reading.asset_code or reading.rfid_code or "-",
                reading.read_method or "-",
                reading.read_at.strftime("%d/%m/%Y %H:%M") if reading.read_at else "-"
            ]
            for col, value in enumerate(row_data, start=1):
                cell = ws_unreg.cell(row=i, column=col, value=value)
                cell.border = thin_border

        ws_unreg.column_dimensions['A'].width = 30
        ws_unreg.column_dimensions['B'].width = 15
        ws_unreg.column_dimensions['C'].width = 18

    # Salvar
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"inventario_{session.code}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
